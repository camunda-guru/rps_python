'''
Created on 02-Jun-2016

@author: BALASUBRAMANIAM
'''

#import required libraries
from openpyxl import load_workbook
 
#read  from excel file
wb = load_workbook('G:/Local Disk/Course content/Net 4 0 Contents Fresher Program-2016.xlsx')
print(type(wb))
print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet5')

print(sheet['D5'].value)

for cell in range(1,sheet.max_row):

    print(cell,sheet.cell(row=cell,column=4).value)
    
    
for r in range(1,17):

    print(r,sheet.cell(row=r,column=3).value,sheet.cell(row=r,column=4).value,sheet.cell(row=r,column=5).value)